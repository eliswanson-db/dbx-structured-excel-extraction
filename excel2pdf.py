# Databricks notebook source
# MAGIC %pip install -qqqq openpyxl==3.1.5 xhtml2pdf==0.2.17

# COMMAND ----------

dbutils.library.restartPython()

# COMMAND ----------

from typing import List, Any
from dataclasses import dataclass
import os
import io
from openpyxl import load_workbook
from xhtml2pdf import pisa
from pyspark.sql import DataFrame
from pyspark.sql.functions import col, md5, concat_ws, pandas_udf
from pyspark.sql.types import StructType, StructField, StringType
import pandas as pd

# COMMAND ----------

dbutils.widgets.text("catalog_name", "")
dbutils.widgets.text("source_schema", "")
dbutils.widgets.text("source_volume", "")
dbutils.widgets.text("dest_schema", "")
dbutils.widgets.text("dest_volume", "")
dbutils.widgets.text("dest_subfolder", "")
dbutils.widgets.text("dest_metadata_table", "")
dbutils.widgets.text("worksheet_name", "Database")

catalog_name = dbutils.widgets.get("catalog_name")
source_schema = dbutils.widgets.get("source_schema")
source_volume = dbutils.widgets.get("source_volume")
dest_schema = dbutils.widgets.get("dest_schema")
dest_volume = dbutils.widgets.get("dest_volume")
dest_subfolder = dbutils.widgets.get("dest_subfolder")
dest_metadata_table = dbutils.widgets.get("dest_metadata_table")
default_worksheet_name = dbutils.widgets.get("worksheet_name")

# Configure table extraction ranges
# Commented out sections you can add if you want additional data sections. 
# Each will show up on another page in the PDF.
table_configs = {
    "Section 1: Basic Information": {
        "min_row": 1,
        "max_row": 18,
        "min_col": 1,
        "max_col": 2,
        "page_break": False,
    },
    "Section 2: Data Section": {
        "min_row": 8,
        "max_row": 50,
        "min_col": 3,
        "max_col": 6,
        "page_break": True,
    },
    # "Section 3: Data Section": {
    #     "min_row": 8,
    #     "max_row": 50,
    #     "min_col": 7,
    #     "max_col": 9,
    #     "page_break": True,
    # },
    # "Section 4: Data Section": {
    #     "min_row": 8,
    #     "max_row": 50,
    #     "min_col": 10,
    #     "max_col": 12,
    #     "page_break": True,
    # },
}

result_schema = StructType([StructField("dest_path", StringType(), True)])

@dataclass
class Config:
    """Configuration class for Excel to PDF processing."""

    catalog_name: str
    source_schema: str
    source_volume: str
    dest_schema: str
    dest_volume: str
    dest_subfolder: str
    dest_metadata_table: str
    worksheet_name: str
    table_configs: dict

    @property
    def source_path(self) -> str:
        """Get the source volume path."""
        return f"/Volumes/{self.catalog_name}/{self.source_schema}/{self.source_volume}"

    @property
    def dest_path(self) -> str:
        """Get the destination volume path."""
        return f"/Volumes/{self.catalog_name}/{self.dest_schema}/{self.dest_volume}/{self.dest_subfolder}"

    @property
    def dest_table(self) -> str:
        """Get the destination table name."""
        return f"{self.catalog_name}.{self.dest_schema}.{self.dest_metadata_table}"

    @property
    def checkpoint_folder(self) -> str:
        """Get the checkpoint folder path."""
        return f"{self.dest_path}/checkpoints"


config = Config(
    catalog_name=catalog_name,
    source_schema=source_schema,
    source_volume=source_volume,
    dest_schema=dest_schema,
    dest_volume=dest_volume,
    dest_subfolder=dest_subfolder,
    dest_metadata_table=dest_metadata_table,
    worksheet_name=default_worksheet_name,
    table_configs=table_configs,
)

# COMMAND ----------


def extract_range(
    ws: Any, min_row: int, max_row: int, min_col: int, max_col: int
) -> List[List[Any]]:
    """
    Extract a range of cells from an Excel worksheet.

    Args:
        ws: The worksheet object
        min_row: Starting row number
        max_row: Ending row number
        min_col: Starting column number
        max_col: Ending column number

    Returns:
        List of lists containing cell values
    """
    return [
        [cell.value for cell in row]
        for row in ws.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
        )
    ]


def html_table(data: List[List[Any]], title: str, page_break: bool = False) -> str:
    """
    Convert data to HTML table format with fixed pixel column widths.

    Args:
        data: List of lists containing table data
        title: Title for the table section
        page_break: Whether to add a page break before this table

    Returns:
        HTML string representation of the table
    """
    if not data:
        return f"<h2>{title}</h2><p>No data</p>\n"

    num_cols = len(data[0]) if data else 0
    table_width = 600  # Fixed total width in pixels
    col_width = table_width // num_cols if num_cols > 0 else table_width

    page_break_style = "page-break-before: always;" if page_break else ""
    html = f"<h2 style='{page_break_style}'>{title}</h2>\n"
    html += f"<table border='1' style='table-layout: fixed; width: {table_width}px; border-collapse: collapse;'>\n"

    for row in data:
        html += "<tr>"
        for cell in row:
            cell_content = str(cell) if cell is not None else ""
            html += f"<td style='width: {col_width}px; max-width: {col_width}px; word-wrap: break-word; padding: 2px 4px; line-height: 1.2; overflow: hidden; vertical-align: top;'>{cell_content}</td>"
        html += "</tr>\n"
    html += "</table>\n"
    return html


def xlsm_to_html(
    xlsm_file: str, worksheet_name: str = "Database", table_configs: dict = None
) -> str:
    """
    Convert Excel file to HTML format by extracting configurable ranges.

    Args:
        xlsm_file: Path to the Excel file
        worksheet_name: Name of the worksheet to extract data from
        table_configs: Dictionary defining table extraction ranges

    Returns:
        HTML string containing formatted tables

    Raises:
        ValueError: If worksheet doesn't exist or file cannot be processed
    """
    try:
        workbook = load_workbook(xlsm_file, read_only=True, data_only=True)

        if worksheet_name not in workbook.sheetnames:
            available_sheets = ", ".join(workbook.sheetnames)
            raise ValueError(
                f"Worksheet '{worksheet_name}' not found. Available sheets: {available_sheets}"
            )

        sheet = workbook[worksheet_name]

        # Use default table config if none provided - single table with 20 rows, 5 columns
        if table_configs is None:
            table_configs = {
                "Data Table": {
                    "min_row": 1,
                    "max_row": 20,
                    "min_col": 1,
                    "max_col": 5,
                    "page_break": False,
                }
            }

        html_content = "<html><head><style>table {border-collapse: collapse; margin: 20px 0;} td {text-align: left;} h2 {margin-top: 30px; margin-bottom: 10px;}</style></head><body>\n"

        # Extract and generate HTML for each configured table
        for title, config in table_configs.items():
            data = extract_range(
                sheet,
                config["min_row"],
                config["max_row"],
                config["min_col"],
                config["max_col"],
            )
            html_content += html_table(
                data, title, page_break=config.get("page_break", False)
            )

        html_content += "</body></html>"

        workbook.close()
        return html_content

    except Exception as e:
        raise ValueError(f"Failed to process Excel file {xlsm_file}: {str(e)}")


def html_to_pdf_bytes(html_content: str) -> bytes:
    """
    Convert HTML content directly to PDF bytes without writing to disk.

    Args:
        html_content: HTML string to convert

    Returns:
        PDF content as bytes
    """
    output_buffer = io.BytesIO()
    pisa_status = pisa.CreatePDF(html_content, dest=output_buffer)

    if pisa_status.err:
        raise RuntimeError("Error converting HTML to PDF")

    return output_buffer.getvalue()


def make_converter_udf(dest_path: str, worksheet_name: str, table_configs: dict):
    """
    Higher-order function returning pandas UDF for Excel to PDF conversion.

    Args:
        dest_path: Destination directory path for PDF files
        worksheet_name: Name of Excel worksheet to process
        table_configs: Dictionary defining table extraction ranges

    Returns:
        Pandas UDF function for converting Excel files to PDF
    """

    @pandas_udf(result_schema)
    def convert_excel_to_pdf(file_paths: pd.Series) -> pd.DataFrame:
        """
        Convert Excel files to PDF format using pandas UDF.

        Args:
            file_paths: Series of file paths to process

        Returns:
            DataFrame with destination paths for generated PDFs
        """
        results = []

        for file_path in file_paths:
            # Generate output filename
            filename = os.path.basename(file_path)
            base_name = os.path.splitext(filename)[0]
            pdf_filename = f"{base_name}.pdf"
            dest_file_path = os.path.join(dest_path, pdf_filename).replace("dbfs:", "")
            source_file_path = file_path.replace("dbfs:", "")

            # Convert Excel to HTML then to PDF
            html_content = xlsm_to_html(source_file_path, worksheet_name, table_configs)
            pdf_bytes = html_to_pdf_bytes(html_content)

            # Write PDF to destination
            with open(dest_file_path, "wb") as f:
                f.write(pdf_bytes)

            results.append({"dest_path": dest_file_path})

        return pd.DataFrame(results)

    return convert_excel_to_pdf


def read_bronze_excel_stream(source_path: str) -> DataFrame:
    """
    Read Excel files from a source volume using structured streaming.

    Args:
        source_path: Path to the source volume containing Excel files

    Returns:
        DataFrame with file metadata for streaming processing
    """
    return (
        spark.readStream.format("cloudFiles")
        .option("cloudFiles.format", "binaryFile")
        .option("cloudFiles.useNotifications", "false")
        .option("cloudFiles.includeExistingFiles", "true")
        .load(source_path)
        .select("modificationTime", "path", "length")
        .filter(
            col("path").rlike(".*\\.(xlsx|xlsm|xls)$")
        )  # If there's a lot of files, this could be moved to a path glob filter for autoloader.
    )


def process_files(config: Config, excel_stream: DataFrame) -> DataFrame:
    """
    Process Excel files by converting them to PDF format.

    Args:
        config: Configuration object with processing parameters
        excel_stream: Stream DataFrame containing Excel file metadata

    Returns:
        DataFrame with processing results and destination paths
    """
    convert_files_udf = make_converter_udf(
        config.dest_path, config.worksheet_name, config.table_configs
    )
    return (
        excel_stream.withColumn("result", convert_files_udf(col("path")))
        .withColumn("dest_path", col("result.dest_path"))
        .withColumn("id", md5(concat_ws("|", col("modificationTime"), col("path"))))
        .select("id", "path", "modificationTime", "dest_path", "length")
    )


def write_stream(processed_stream: DataFrame, config: Config) -> None:
    """
    Write processed stream to the destination table.

    Args:
        processed_stream: DataFrame with processed file information
        config: Configuration object with destination settings
    """
    query = (
        processed_stream.writeStream.option(
            "checkpointLocation", config.checkpoint_folder
        )
        .outputMode("append")
        .trigger(availableNow=True)
        .option("mergeSchema", "true")  # Handle schema evolution
        .toTable(config.dest_table)
    )
    query.awaitTermination()


def setup_environment(config: Config) -> None:
    """
    Set up the processing environment by creating necessary volumes and directories.

    Args:
        config: Configuration object with setup parameters
    """
    try:
        # Create destination volume if it doesn't exist
        spark.sql(
            f"CREATE VOLUME IF NOT EXISTS {config.catalog_name}.{config.dest_schema}.{config.dest_volume}"
        )

        # Create destination table if it doesn't exist
        spark.sql(
            f"""
            CREATE TABLE IF NOT EXISTS {config.dest_table} (
                id STRING,
                path STRING,
                modificationTime TIMESTAMP,
                dest_path STRING,
                length BIGINT
            ) USING DELTA
        """
        )

        # Create destination directories
        try:
            os.makedirs(config.dest_path)
            os.makedirs(config.checkpoint_folder)
        except FileExistsError:
            print("Folder already exists...")
        except Exception as e:
            print(f"{e}")

        print("Environment setup completed successfully")
        print(
            f"- Destination volume: {config.catalog_name}.{config.dest_schema}.{config.dest_volume}"
        )
        print(f"- Destination table: {config.dest_table}")
        print(f"- PDF output directory: {config.dest_path}")

    except Exception as e:
        raise RuntimeError(f"Failed to setup environment: {str(e)}")


def main(config: Config) -> None:
    """
    Main processing function to orchestrate Excel to PDF conversion.

    Args:
        config: Configuration object with all processing parameters
    """
    setup_environment(config)
    excel_stream = read_bronze_excel_stream(config.source_path)
    processed_stream = process_files(config, excel_stream)
    write_stream(processed_stream, config)


# COMMAND ----------


def validate_configuration(config: Config) -> None:
    """
    Validate that all required configuration parameters are provided.

    Args:
        config: Configuration object to validate

    Raises:
        ValueError: If any required parameter is missing or invalid
    """
    if not config.catalog_name:
        raise ValueError("catalog_name is required")
    if not config.source_schema:
        raise ValueError("source_schema is required")
    if not config.source_volume:
        raise ValueError("source_volume is required")
    if not config.dest_schema:
        raise ValueError("dest_schema is required")
    if not config.dest_volume:
        raise ValueError("dest_volume is required")
    if not config.dest_metadata_table:
        raise ValueError("dest_metadata_table is required")
    if not config.worksheet_name:
        raise ValueError("worksheet_name is required")


# COMMAND ----------

# Execute the Excel to PDF conversion pipeline
try:
    validate_configuration(config)
    print("Starting Excel to PDF conversion pipeline...")
    print(f"Source: {config.source_path}")
    print(f"Destination: {config.dest_path}")
    print(f"Target table: {config.dest_table}")

    main(config)
    print("Pipeline completed successfully!")

except Exception as e:
    print(f"Pipeline failed with error: {e}")
    raise

# COMMAND ----------
