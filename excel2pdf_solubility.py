# Databricks notebook source
# MAGIC %pip install -qqqq openpyxl==3.1.5 xhtml2pdf==0.2.17

# COMMAND ----------

dbutils.library.restartPython()

# COMMAND ----------

dbutils.widgets.text("catalog_name", "")
dbutils.widgets.text("source_schema", "")
dbutils.widgets.text("source_volume", "")
dbutils.widgets.text("dest_schema", "")
dbutils.widgets.text("dest_volume", "")
dbutils.widgets.text("dest_subfolder", "")
dbutils.widgets.text("dest_metadata_table", "")

catalog_name = dbutils.widgets.get("catalog_name")
source_schema = dbutils.widgets.get("source_schema")
source_volume = dbutils.widgets.get("source_volume")
dest_schema = dbutils.widgets.get("dest_schema")
dest_volume = dbutils.widgets.get("dest_volume")
dest_subfolder = dbutils.widgets.get("dest_subfolder")
dest_metadata_table = dbutils.widgets.get("dest_metadata_table")

config_map = {
    "catalog_name": catalog_name,
    "source_schema": source_schema,
    "source_volume": source_volume,
    "dest_schema": dest_schema,
    "dest_volume": dest_volume,
    "dest_subfolder": dest_subfolder,
    "dest_metadata_table": dest_metadata_table
}

# COMMAND ----------

workbook_name = "/Volumes/dbxmetagen/default/pfizer_xl/003c5881-8b5e-4f76-b0a5-55305d9f38fd_00705805-0802 PF-07104665 Solubility Screen NEAT_UV.xlsm"

# COMMAND ----------

from openpyxl import load_workbook
import os
from xhtml2pdf import pisa

def extract_range(ws, min_row, max_row, min_col, max_col):
    return [
        [cell.value for cell in row]
        for row in ws.iter_rows(
            min_row=min_row, max_row=max_row,
            min_col=min_col, max_col=max_col
        )
    ]

def html_table(data, title):
    html = f"<h2>{title}</h2><table border='1'>\n"
    for row in data:
        html += "<tr>" + ''.join(f"<td>{cell if cell is not None else ''}</td>" for cell in row) + "</tr>\n"
    html += "</table>\n"
    return html

def xlsm_to_html(xlsm_file):
    workbook = load_workbook(xlsm_file)
    sheet = workbook['Database']
    data1 = extract_range(sheet, 1, 18, 1, 2)
    data2 = extract_range(sheet, 8, 50, 3, 9)
    #data3 = extract_range(sheet, 8, 50, 7, 9)

def define_html_body(data_arr: list[list[str]])
    html_content = "<html><body>\n"
    html_content += html_table(data1, "Section 1: A1:B18")
    html_content += html_table(data2, "Section 2: E3:J50")
    #html_content += html_table(data3, "Section 3: H3:J50")
    html_content += "</body></html>"
    return html_content

def create_path(dest_path):
    if not os.path.exists(dest_path):
        os.mkdir(dest_path)
        print("Created path...")
    if not os.path.exists(dest_path):
        os.mkdir(dest_path)
        print("Path exists...")

def write_html()
    # TODO: why are we writing out the html content at all? Do we need that, or could we just define the html and go strait to source html in html_file_to_pdf? I think so.
    html_content = define_html_body()
    with open("output.html", "w", encoding="utf-8") as f:
        f.write(html_content)

def html_file_to_pdf(html_file, pdf_file):
    with open(html_file, "r", encoding="utf-8") as f:
        source_html = f.read()
    with open(pdf_file, "wb") as output:
        pisa_status = pisa.CreatePDF(source_html, dest=output)
    return pisa_status.err

def make_converter_udf():
    """Higher-order function returning pandas UDF to handle passing config."""
    @pandas_udf(profile_schema)
    def convert_files():
        pass
        # TODO: UDF should call functions to load excel based on path, convert to html tables, then convert to pdf.
        # Please use higher-order function approach to improve serializability and pass variables if needed.
        
def read_bronze_pdf_stream(source_path: str) -> DataFrame:
    """
    Read the bronze PDF stream.
    """
    return (spark.readStream
         .format("cloudFiles")
         .option("cloudFiles.format", "binaryFile")
         .load(source_path)
         .select("modificationTime", "path", "length")
         )

def process_files(profiler_config: ProfilerConfig, bronze_pdf_stream: DataFrame) -> DataFrame:
    """
    Process the files.
    """
        convert_files_udf = make_converter_udf()
        return (bronze_pdf_stream
                .withColumn("dest_path", convert_files_udf(bronze_pdf_stream["path"]))
                .withColumn("id", md5(concat_ws("|", col("modificationTime"), col("path"))))
                .select("id", "path", "modificationTime", "dest_path")
        )

def write_stream(processed_stream: DataFrame, config: ProfilerConfig) -> DataFrame:
    return (
        processed_stream.writeStream
            .option("checkpointLocation", config.checkpoint_folder)
            .outputMode("append")
            .trigger(availableNow=True)
            .toTable(config.dest_table)
            .awaitTermination()
        )

def create_subfolder(config_map: Dict[str, Union[str, int]]) -> None:
    """
    Create a subfolder if it doesn't exist.
    """
    try:
        os.mkdir(config_map.dest_path)
    except FileExistsError:
        print("Folder already exists. Skipping make directory...")
    except Exception as e:
        print(f"{e}")

def main(config_map):
    spark.sql(f"CREATE VOLUME IF NOT EXISTS {config.get('dest_volume')}")
    create_subfolder(config_map.get('dest_path'))
    bronze_pdf_stream = read_bronze_pdf_stream(config.get('source_path'))
    processed_stream = process_files(config, bronze_pdf_stream)
    write_stream(processed_stream, config)


# COMMAND ----------

# TODO: add a sanity check here
